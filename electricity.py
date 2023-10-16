from openpyxl import Workbook
wb = Workbook()
        
# grab the active worksheet
ws = wb.active
ws.title = 'Customer'

# Data can be assigned directly to cells
ws['A1'] = 'customer'
ws['B1'] = 'account_no'
ws['C1'] = 'location_coordinates'
ws['D1'] = 'status'
ws['E1'] = 'phone_number'

ws['A2'] = 'Bena'
ws['B2'] = '04244389'
ws['C2'] = '0.1234,1.5678'
ws['D2'] = 'Active'
ws['E2'] = '0703533798'

ws['A3'] = 'Bridget'
ws['B3'] = '9076453'
ws['C3'] = '70.6745, 20.5689'
ws['D3'] = 'Active'
ws['E3'] = '0702426585'

ws['A4'] = 'Daphne'
ws['B4'] = '0678254'
ws['C4'] = '0.4678,2.4678'
ws['D4'] = 'Active'
ws['E4'] = '0789125096'

cell_range1 = ws['A2':'E2']
for row in cell_range1:
  for cell in row:
    print(f'{cell}: {cell.value}')
    
ws1 = wb.create_sheet("Pricing")       
ws1['A1'] = 'flat_units'
ws1['B1'] = 'flat_unit_rate'
ws1['C1'] = 'other_unit_rate'
ws1['D1'] = 'vate_rate'
ws1['E1'] = 'service_fee'

ws1['A2'] = 15
ws1['B2'] = 150
ws1['C2'] = 520.6
ws1['D2'] = 18
ws1['E2'] = 3360

cell_range2 = ws1['A2':'E2'] 
for row in cell_range2:  
  for cell in row:
    print(f'{cell}: {cell.value}')
    
ws2 = wb.create_sheet("Billing")     
ws2['A1'] = 'customer_id'
ws2['B1'] = 'current_reading'
ws2['C1'] = 'previous_reading'
ws2['D1'] = 'balance_bf'
ws2['E1'] = 'bill_amount'

ws2['A2'] = 1
ws2['B2'] = 37.432
ws2['C2'] = 0
ws2['D2'] = 0

ws2['A3'] = 2
ws2['B3']= 30
ws2['C3'] = 10
ws2['D3'] = 10

ws2['A4'] = 3
ws2['B4'] = 40
ws2['C4'] = 0
ws2['D4'] = 20

cell_range3 = ws2['A2':'E2'] 
for row in cell_range3:
  for cell in row:
    print(f'{cell}: {cell.value}')
    
# Save the file
wb.save("week1/bills/electricity.xlsx")

